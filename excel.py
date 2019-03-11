# -*- coding: utf-8 -*-
from openpyxl import Workbook

wb = Workbook()

# grab the active worksheet
ws = wb.active

ws.title = 'new title'
ws.sheet_properties.tabColor = "1072BA"

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")

